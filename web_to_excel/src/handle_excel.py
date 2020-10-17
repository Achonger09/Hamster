
from openpyxl import load_workbook
from openpyxl.styles import colors, Alignment,Border, Side, PatternFill, Font, GradientFill, Alignment
import time

'''
path = "./excel/source.xlsx"
sheet_name = "Sheet1"
timestr = time.strftime("%d-%H-%M-%S")
des_path = "./excel/result%s.xlsx" % timestr
key_line = 1  ## 关键字行号
pipe_names = ["数据2"]  ## 流水线name列表
pipe_name_row = 'A' ## 流水线名字列号
'''
class HandleExcel(object):

    def __init__(self,source_path,des_path = None,config=None):
        '''
        dource_path str 源excel路径
        des_path str  输出excel路径
        config Config 配置文件
        '''
        ## open excel
        self.source_wb = load_workbook(source_path)
        self.ws = self.source_wb.active
        print(self.ws.max_row)
        print(self.ws.max_column)
        self.key_line = 1
        self.pipe_name_row = 'A'
        print("load excel %s success" % source_path)
        #self.bold_itatic_24_font = Font(name='等线', size=24, italic=True, color=colors.RED, bold=True)

    
    def __find_row_by_name(self,line, name):
        '''
        line 
        name 
        从列中选择对应name的行号
        '''
        print("__find_row_by_name(%s, %s) failed" %(str(line),str(name)))
        for col in self.ws[line]:
            if col.value == name:
                print("__find_row_by_name(%s, %s) %s" %(str(line),str(name),str(col.row)))
                return col.column
        print("__find_row_by_name(%s, %s) failed" %(str(line),str(name)))
        return None
    
    def __find_col_by_name(self,row, name):
        '''
        row
        name
        从行中选择对应name的列号
        '''
        print("__find_col_by_name(%s, %s) failed" %(str(row),str(name)))
        for row in self.ws[row]:
            print(row.value , name)
            if row.value == name:
                print("__find_col_by_name(%s, %s) %s" %(str(row),str(name),str(row.column)))
                return row.row
        print("__find_col_by_name(%s, %s) failed" %(str(row),str(name)))

    def write_by_name(self,pipe_name, col_name, values):
        '''
        pipe_name 流水线名称
        col_name 指标名称
        '''
        row = self.__find_col_by_name(self.pipe_name_row, pipe_name)
        col = self.__find_row_by_name(self.key_line, col_name)
        print("write %s%s : %s" %(str(row),str(col),values))
        self.ws["%s%s" %(str(col),str(row))].value = values
        #self.ws["%s%s" %(str(col),str(row))].font = self.bold_itatic_24_font
        

    def save(self,path):
        '''
        保存数据到文件path
        '''
        self.source_wb.save(path)
'''
he = HandleExcel(path)
he.write_by_name("数据3","标题4","test")
he.save(des_path)
'''